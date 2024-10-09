import os
from typing import Callable, List, Tuple, Union

import numpy as np
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

##################################################
# Variable globale

DARK_BLUE = "9683EC"
LIGHT_BLUE = "E6E6FA"

##################################################
# Fonction d'enregistrement et de sauvegarde


def load_pbi(file: str) -> pd.DataFrame:
    return pd.read_csv(file, sep=";", low_memory=False)


def create_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def save_workbook(wb: Workbook, output: str) -> None:
    wb.save(output)


##################################################
# Fonction de style excel


def adjust_style(ws: Worksheet) -> None:
    # Couleurs

    dark_blue_fill = PatternFill(
        start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid"
    )
    light_blue_fill = PatternFill(
        start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    thick_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = thin_border

    # Changer la couleur de fond de la colonne A en bleu clair
    for cell in ws["A"]:
        cell.fill = light_blue_fill
        cell.border = thick_border

    # Changer la couleur de fond des lignes 1 et 2
    for row in [1, 2]:
        for cell in ws[row]:
            cell.fill = light_blue_fill
            cell.font = Font(bold=True)
            cell.border = thick_border

    # Parcourir les lignes pour trouver la valeur "Total" dans la colonne A
    for row in ws.iter_rows(min_row=3):
        if row[0].value == "Total":
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = dark_blue_fill
                cell.border = thick_border


def add_headers1(ws: Worksheet, bo: bool) -> None:
    name = None
    if bo:
        name = "6 mois"
    else:
        name = "Situation actuelle"
    headers = [
        [
            name,
            "Nombre total de titulaires de la certification",
            "Nombre de titulaires après neutralisation",
            "Répondants",
            "",
            "Répondants après neutralisation",
            "",
            "En activité professionnelle toutes fonctions",
            "",
            "Nombre de titulaires exerçant principalement les activités visées par la certification",
            "",
            "Rémunération brute annuelle moyenne des titulaires exerçant les activités visées",
        ],
        ["Promo", "", "", "Nbre", "%", "Nbre", "%", "Nbre", "%", "Nbre", "%", ""],
    ]

    for row_num, header_row in enumerate(headers, start=1):
        for col_num, header in enumerate(header_row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            if (row_num <= 2) | (col_num == 1):
                cell.fill = PatternFill(
                    start_color="A9C6FF", end_color="A9C6FF", fill_type="solid"
                )


def adjust_cell1(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 20

    ws.row_dimensions[1].height = 75

    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:E1")
    ws.merge_cells("F1:G1")
    ws.merge_cells("H1:I1")
    ws.merge_cells("J1:K1")
    ws.merge_cells("L1:L2")


def add_headers2(ws: Worksheet) -> None:
    headers = [
        [
            "Promo",
            "Formation Initiale",
            "",
            "Formation Continue",
            "",
            "VAE",
            "Candidatures libres (Hors parcours)",
            "Nombre Total de titres",
        ],
        [
            "",
            "Statut d'élève ou d'étudiant",
            "En contrat d'apprentissage",
            "Statut de stagiaire de la formation professionnelle",
            "Contrat de professionnalisation",
            "",
            "",
            "",
        ],
    ]

    for row_num, header_row in enumerate(headers, start=1):
        for col_num, header in enumerate(header_row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if (row_num <= 2) | (col_num == 1):
                cell.fill = PatternFill(
                    start_color="A9C6FF", end_color="A9C6FF", fill_type="solid"
                )


def adjust_cell2(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20

    ws.row_dimensions[2].height = 70

    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:C1")
    ws.merge_cells("D1:E1")
    ws.merge_cells("F1:F2")
    ws.merge_cells("G1:G2")
    ws.merge_cells("H1:H2")


##################################################
# Fonction utilitaire


def percent(up: int, down: int) -> str:
    if down == 0:
        return "0%"
    return str(int(round((up / down) * 100))) + "%"


def salary(
    sal: Union[float, int],
) -> str:
    return str(int(round(sal))) + "€"


def merge_all(*df: pd.DataFrame) -> pd.DataFrame:
    if not df:
        raise ValueError("At least one DataFrame is required")

    data = df[0]

    for _df_ in df[1:]:
        data = pd.merge(data, _df_, on="Promotion", how="outer")

    return data


def count_colonne(
    df1: pd.DataFrame, df2: pd.DataFrame, col1: str, col2: str, lst_prom: pd.DataFrame
) -> pd.DataFrame:
    if df1.empty:
        ret = pd.DataFrame()
        ret["Promotion"] = lst_prom["Promotion"]
        ret[col1] = np.nan
        ret[col1 + " %"] = np.nan
    else:
        ret = df1.groupby("Promotion").size().reset_index(name=col1)
        ret[col1 + " %"] = ret.apply(
            lambda row: f"{round( row[col1] / df2.loc[df2['Promotion'] == row['Promotion'], col2].values[0] * 100)}%",
            axis=1,
        )

    ret[col1] = ret[col1].fillna(0)
    ret[col1 + " %"] = ret[col1 + " %"].fillna("0%")

    return ret


def path(src: str) -> str:
    normalized_path = os.path.normpath(src)

    return normalized_path.replace("\\", "/")


##################################################
# Fonction obtention data


def data_salary(
    df: pd.DataFrame, salaire_min: int, salaire_max: int, update: Callable
) -> pd.DataFrame:
    count = df.groupby("Promotion").size().reset_index(name="Count")
    lst_prom = pd.DataFrame()
    lst_prom["Promotion"] = count["Promotion"]
    non_neutr_data = df[df["Neutralisation"] != "neutralisés"]
    non_neutr_count = (
        non_neutr_data.groupby("Promotion").size().reset_index(name="Non neutralisés")
    )
    update()
    rep_data = df[df["Répondant"] == 1]
    rep_count = count_colonne(rep_data, count, "Répondant", "Count", lst_prom)

    rep_non_neutr_data = non_neutr_data[non_neutr_data["Répondant"] == 1]
    rep_non_neutr_count = count_colonne(
        rep_non_neutr_data,
        non_neutr_count,
        "Répondant non neutralisés",
        "Non neutralisés",
        lst_prom,
    )

    emploi_data = rep_non_neutr_data[
        rep_non_neutr_data["Situation professionnelle niv2"] == "En emploi"
    ]
    emploi_count = count_colonne(
        emploi_data,
        rep_non_neutr_count,
        "En emploi",
        "Répondant non neutralisés",
        lst_prom,
    )

    fv_data = rep_non_neutr_data[
        rep_non_neutr_data["Situation professionnelle niv3"] == "FV"
    ]
    fv_count = count_colonne(
        fv_data, rep_non_neutr_count, "FV", "Répondant non neutralisés", lst_prom
    )

    salaire_data = fv_data[
        (fv_data["Salaire brut global annuel"] > salaire_min)
        & (fv_data["Salaire brut global annuel"] < salaire_max)
    ]
    salaire_avg = (
        salaire_data.groupby("Promotion")
        .agg(
            {
                "Salaire brut global annuel": lambda x: (
                    salary(np.nanmean(x))
                    if pd.notna(np.nanmean(x))
                    else "Non renseigné"
                )
            }
        )
        .reset_index()
    )
    salaire_avg.rename(
        columns={"Salaire brut global annuel": "Salaire moyen"}, inplace=True
    )

    data_ret = merge_all(
        count,
        non_neutr_count,
        rep_count,
        rep_non_neutr_count,
        emploi_count,
        fv_count,
        salaire_avg,
    )
    data_ret["Count"] = data_ret["Count"].fillna(0)
    data_ret["Non neutralisés"] = data_ret["Non neutralisés"].fillna(0)
    data_ret["Salaire moyen"] = data_ret["Salaire moyen"].fillna("Non renseigné")
    update()
    data_ret = data_ret.sort_values(by="Promotion", ascending=True)

    totals = data_ret.drop(columns=["Promotion"]).select_dtypes(include=np.number).sum()

    moyenne_salaire = pd.to_numeric(
        salaire_data["Salaire brut global annuel"], errors="coerce"
    ).mean()
    if not pd.api.types.is_numeric_dtype(moyenne_salaire):
        moyenne_salaire = "Non renseigné"
    else:
        moyenne_salaire = salary(moyenne_salaire)

    total_row = pd.DataFrame(totals).transpose()
    total_row["Promotion"] = "Total"
    total_row["Répondant %"] = percent(
        total_row["Répondant"].item(), total_row["Count"].item()
    )
    total_row["Répondant non neutralisés %"] = percent(
        total_row["Répondant non neutralisés"].item(),
        total_row["Non neutralisés"].item(),
    )
    total_row["En emploi %"] = percent(
        total_row["En emploi"].item(), total_row["Répondant non neutralisés"].item()
    )
    total_row["FV %"] = percent(
        total_row["FV"].item(), total_row["Répondant non neutralisés"].item()
    )
    total_row["Salaire moyen"] = moyenne_salaire

    data_ret = pd.concat([data_ret, total_row], ignore_index=True)
    update()
    return data_ret


def data_va(df: pd.DataFrame, update: Callable) -> pd.DataFrame:
    eleve = (
        df[df["Voie d'accès"] == "Etudiant (Formation initiale)"]
        .groupby("Promotion")
        .size()
        .reset_index(name="Élève")
    )
    apprenti = (
        df[df["Voie d'accès"] == "Apprenti (Contrat d’apprentissage)"]
        .groupby("Promotion")
        .size()
        .reset_index(name="Apprenti")
    )
    stage = (
        df[
            df["Voie d'accès"]
            == "Stagiaire de la formation professionnelle (Formation continue)"
        ]
        .groupby("Promotion")
        .size()
        .reset_index(name="Stage")
    )
    pro = (
        df[df["Voie d'accès"] == "En contrat de professionnalisation"]
        .groupby("Promotion")
        .size()
        .reset_index(name="Pro")
    )
    vae = (
        df[df["Voie d'accès"] == "VAE"]
        .groupby("Promotion")
        .size()
        .reset_index(name="VAE")
    )
    hors = (
        df[df["Voie d'accès"] == "Candidature libre (Hors parcours)"]
        .groupby("Promotion")
        .size()
        .reset_index(name="Hors")
    )
    count = df.groupby("Promotion").size().reset_index(name="Total")

    data = merge_all(eleve, apprenti, stage, pro, vae, hors, count)

    update()
    data = data.sort_values(by="Promotion", ascending=True)

    total = data.drop(columns=["Promotion"]).select_dtypes(include=np.number).sum()
    total = pd.DataFrame(total).transpose()

    total["Promotion"] = "Total"
    update()
    data = pd.concat([data, total], ignore_index=True)

    data = data.fillna(0)

    return data


##################################################
# Fonction utilitaire app


def filtered_tab1(df: pd.DataFrame) -> pd.DataFrame:
    return df[df["Situation questionnaire"] == "6 mois"]


def filtered_tab2(df: pd.DataFrame) -> pd.DataFrame:
    return df[df["Situation questionnaire"] == "situation acctuelle"]


def filtered_tab3(df: pd.DataFrame) -> pd.DataFrame:
    return df[df["Situation questionnaire"] == "6 mois"]


def filtered_promo(df: pd.DataFrame, promo: List[int]) -> pd.DataFrame:
    return df[df["Promotion"].isin(promo)]


def filtered_promo_true(
    df: pd.DataFrame, promo: List[Tuple[int, bool]]
) -> pd.DataFrame:
    promo_lst = [prom for (prom, val) in promo if val]
    return filtered_promo(df, promo_lst)


def filtered_ecole(df: pd.DataFrame, ecole: List[str]) -> pd.DataFrame:
    return df[df["Ecole (UC)"].isin(ecole)]


def filtered_ecole_true(
    df: pd.DataFrame, ecole: List[Tuple[str, bool]]
) -> pd.DataFrame:
    ecole_lst = [eco for (eco, val) in ecole if val]
    return filtered_ecole(df, ecole_lst)


def filtered_all_unique(df: pd.DataFrame) -> pd.DataFrame:
    return df["UNIQUE_KEY"].unique()


def get_promo_lst(df: pd.DataFrame) -> List[int]:
    return df["Promotion"].unique().tolist()


def get_ecole_lst(df: pd.DataFrame) -> List[str]:
    return df["Ecole (UC)"].unique().tolist()


def check_pbifile(file: str) -> bool:
    df = pd.read_csv(file, sep=";", nrows=0)
    col = df.columns.tolist()

    columns = [
        "Prénom",
        "Nom",
        "Email",
        "Ecole (UC)",
        "Langue",
        "Nom campagne",
        "ID Formulaire",
        "Nom Formulaire",
        "Réponse formulaire URL",
        "Certification code",
        "Titre/diplôme",
        "Projet de découpage",
        "ID_alumni",
        "Statut Désinscrit",
        "Date de Jury",
        "Situation questionnaire",
        "Durée post-diplôme",
        "Voie d'accès",
        "Voie accès cocktail",
        "Cursus",
        "Principale expérience avant la formation",
        "Durée principale expérience avant la formation",
        "UNIQUE_KEY",
        "Diplôme avant la formation",
        "Bourse avant la formation",
        "Intitulé certification avant la formation",
        "Spécialité avant la formation",
        "Situation professionnelle avant la formation",
        "NPS insertion",
        "Commentaire général 1",
        "Commentaire général 2",
        "Mention au bac",
        "Ecole actuelle",
        "Téléphone",
        "Promotion",
        "Promotion Date",
        "Titre/diplôme.1",
        "Spécialité actuelle",
        "Fonction",
        "Fonction - autre",
        "Fonction visée",
        "Question Fonction2",
        "Fonction 2",
        "Fonction - autre 2",
        "Fonction visée 2",
        "HFV choisi 2",
        "Situation professionnelle choix",
        "En emploi",
        "Situation étudiant",
        "Situation professionnelle hors en emploi",
        "Situation professionnelle",
        "Région/Pays",
        "Région/Pays - autre",
        "Statut",
        "Contrat",
        "Cible",
        "HFV choisi",
        "Salaire brut annuel",
        "Salaire brut annuel 2 ",
        "Primes brutes annuelles",
        "Fourchette du salaire brut annuel",
        "Fourchette du salaire brut annuel_2",
        "Salaire brut global annuel",
        "Avantages natures",
        "Gestion budget",
        "Souhait d'investissement",
        "Souhait d'investissement - autre",
        "Temps complet",
        "Nombre d'heures par semaine",
        "Temps complet choix",
        "Neutralisation",
        "Inactif",
        "Secteur d'activité",
        "Nom employeur",
        "Encadrement",
        "Contrat niv1",
        "Contrat niv2",
        "Contrat niv3",
        "Situation professionnelle niv1",
        "Situation professionnelle niv2",
        "Situation professionnelle niv3",
        "Situation professionnelle niv4",
        "Statut niv1",
        "Statut niv2",
        "ID_réponse",
        "Répondant",
        "Taille organisation",
        "Nouvelles competences",
        "Compétence_1",
        "Compétence_2",
        "Compétence_3",
        "Compétence_4",
        "Compétence_5",
        "Compétence_6",
        "Compétence_7",
        "Compétence_8",
        "Compétence_9",
        "Compétence_10",
        "Compétence_11",
        "Compétence_12",
        "Compétence_13",
        "Compétence_14",
    ]

    if len(col) != len(columns):
        return False

    for i in range(len(col)):
        if columns[i] != col[i]:
            return False
    return True


def create_lst_bool_value_int(lst: List[Union[int, str]]) -> List[Tuple[int, bool]]:
    ret = [(item, True) for item in lst]
    return sorted(ret, key=lambda x: x[0])


def create_lst_bool_value_str(lst: List[Union[int, str]]) -> List[Tuple[str, bool]]:
    ret = [(item, True) for item in lst]
    return ret


##################################################
# Fonction création tableau


def create_table_salary(
    ws: Worksheet,
    df: pd.DataFrame,
    salaire_min: int,
    salaire_max: int,
    update: Callable,
    tab: bool,
) -> None:
    add_headers1(ws, tab)
    data = data_salary(df, salaire_min, salaire_max, update)
    for idx, row in data.iterrows():
        ws.append(
            [
                row["Promotion"],
                row["Count"],
                row["Non neutralisés"],
                row["Répondant"],
                row["Répondant %"],
                row["Répondant non neutralisés"],
                row["Répondant non neutralisés %"],
                row["En emploi"],
                row["En emploi %"],
                row["FV"],
                row["FV %"],
                row["Salaire moyen"],
            ]
        )

    update()
    adjust_style(ws)
    adjust_cell1(ws)


def create_table_va(ws: Worksheet, df: pd.DataFrame, update: Callable) -> None:
    add_headers2(ws)

    data = data_va(df, update)

    for idx, row in data.iterrows():
        ws.append(
            [
                row["Promotion"],
                row["Élève"],
                row["Apprenti"],
                row["Stage"],
                row["Pro"],
                row["VAE"],
                row["Hors"],
                row["Total"],
            ]
        )

    update()
    adjust_style(ws)
    adjust_cell2(ws)


def create_table_src(ws: Worksheet, df: pd.DataFrame, update: Callable) -> None:
    dark_blue_fill = PatternFill(
        start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid"
    )
    light_blue_fill = PatternFill(
        start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    thick_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    update()

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    update()

    for cell in ws[1]:
        cell.fill = light_blue_fill

        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thick_border

    update()

    color_bool = False

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if color_bool:
                cell.fill = light_blue_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        color_bool = not color_bool

    update()

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    update()


def create_tab1(
    wb: Workbook, pbi: pd.DataFrame, min: int, max: int, update: Callable
) -> None:
    ws = wb.create_sheet()
    ws.title = "6 mois"
    update()
    create_table_salary(ws, pbi, min, max, update, True)


def create_tab2(
    wb: Workbook, pbi: pd.DataFrame, min: int, max: int, update: Callable
) -> None:
    ws = wb.create_sheet()
    ws.title = "Situation actuelle"
    update()
    create_table_salary(ws, pbi, min, max, update, False)


def create_tab3(wb: Workbook, pbi: pd.DataFrame, update: Callable) -> None:
    ws = wb.create_sheet()
    ws.title = "Voie d'accès"
    update()
    create_table_va(ws, pbi, update)


def create_tab4(wb: Workbook, pbi: pd.DataFrame, update: Callable) -> None:
    ws = wb.create_sheet()
    ws.title = "Source"
    update()
    create_table_src(ws, pbi, update)


##################################################
# Fonction total


def fake_update() -> None:
    pass


def tcd(
    input_pbi: pd.DataFrame,
    tab1: bool = True,
    min1: int = 5000,
    max1: int = 150000,
    promo1: List[int] = [],
    ecole1: List[str] = [],
    tab2: bool = True,
    min2: int = 5000,
    max2: int = 150000,
    promo2: List[int] = [],
    ecole2: List[str] = [],
    tab3: bool = True,
    promo3: List[int] = [],
    ecole3: List[str] = [],
    tab4: bool = True,
    update: Callable = fake_update,
) -> Workbook:
    wb = create_workbook()

    pbi = input_pbi

    if tab1:
        df1 = filtered_tab1(pbi)

        if len(promo1) != 0:
            df1 = filtered_promo(df1, promo1)

        if len(ecole1) != 0:
            df1 = filtered_ecole(df1, ecole1)

        create_tab1(wb, df1, min1, max1, update)

    if tab2:
        df2 = filtered_tab2(pbi)

        if len(promo2) != 0:
            df2 = filtered_promo(df2, promo2)

        if len(ecole2) != 0:
            df2 = filtered_ecole(df2, ecole2)

        create_tab2(wb, df2, min2, max2, update)

    if tab3:
        df3 = filtered_tab3(pbi)

        if len(promo3) != 0:
            df3 = filtered_promo(df3, promo3)

        if len(ecole3) != 0:
            df3 = filtered_ecole(df3, ecole3)

        create_tab3(wb, df3, update)

    if tab4:
        create_tab4(wb, pbi, update)

    return wb
